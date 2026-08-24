#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera data/skeleton.json a partir da planilha PokeAgenda (aba 1).
Generates data/skeleton.json from the PokeAgenda spreadsheet (sheet 1).

O "esqueleto" contem apenas dados publicos (entradas, nomes, regioes, flags,
datas de estreia). NENHUMA marca pessoal de captura e exportada.
The "skeleton" holds public data only. NO personal catch marks are exported.

Uso / Usage:
    python tools/export_skeleton.py [caminho/para/PokeAgenda.xlsx]
"""

import json
import os
import sys
import unicodedata
from datetime import date, datetime

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl nao encontrado. Rode: pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DEFAULT_XLSX = os.path.abspath(
    os.path.join(ROOT, "..", "PokéAgenda 2026.xlsx")
)
OUT_PATH = os.path.join(ROOT, "data", "skeleton.json")
SHEET = "PokéAgenda"

# ---------------------------------------------------------------- column map
# Cabecalhos esperados na aba 1 -> chave interna.
# Expected sheet-1 headers -> internal key.
COLS = {
    "Número": "num",
    "Nome": "namePt",
    "Registro": "caught",
    "M": "m",
    "N": "n",
    "F": "f",
    "Brilhante": "shiny",
    "Sombroso": "shadow",
    "Purificado": "purified",
    "Sombroso brilhante": "shadowShiny",
    "Dinamax": "dmax",
    "Dinamax brilhante": "dmaxShiny",
    "Dex brilhante": "shinyDex",
    "XXS": "xxs",
    "XXL": "xxl",
    "Sortudo": "lucky",
    "Perfeito": "perfect",
    "Estreia": "d_base",
    "Estreia brilhante": "d_shiny",
    "Estreia sombroso": "d_shadow",
    "Estreia sombroso brilhante": "d_shadowShiny",
    "Estreia dinamax": "d_dmax",
    "Estreia dinamax brilhante": "d_dmaxShiny",
    "Dex": "dexCanon",
    "Espécie": "speciesPt",
    "Variação regional": "regFormPt",
    "Forma alternativa": "altFormPt",
    "Traje": "costumePt",
    "Species": "speciesEn",
    "Name": "nameEn",
    "Regional Form": "regFormEn",
    "Alternate Form": "altFormEn",
    "Costume": "costumeEn",
    "Região": "region",
    "Megaevolução": "flag_mega",
    "Gigamax": "flag_gmax",
    "Não trocável": "flag_notrade",
    "Exclusivo": "flag_exclusive",
    "Regional": "flag_regional",
    "DiF. Sex.": "gender",
    "Lendário": "flag_legendary",
    "Mítico": "flag_mythical",
    "Ultracriatura": "flag_ultrabeast",
    "Paradoxo": "flag_paradox",
    "Pseudo-lendário": "flag_pseudo",
    "Inicial": "flag_starter",
    "Bebê": "flag_baby",
    "SomenteM": "flag_maleOnly",
    "SomenteF": "flag_femaleOnly",
    "SomenteN": "flag_genderless",
    "ID": "id",
}

# Marcas pessoais: lidas pelo navegador, nunca exportadas aqui.
# Personal marks: read in-browser, never exported here.
MARK_KEYS = [
    "caught", "m", "n", "f", "shiny", "shadow", "purified", "shadowShiny",
    "dmax", "dmaxShiny", "shinyDex", "xxs", "xxl", "lucky", "perfect",
    # living dex (na caixa agora, nao so registrado)
    "living", "livingShiny", "livingShadow", "livingPurified",
    "livingDmax", "livingLucky",
]
DATE_KEYS = ["base", "shiny", "shadow", "shadowShiny", "dmax", "dmaxShiny"]


def norm(s):
    """Compara cabecalhos ignorando acento/caixa/espaco. Header compare."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.lower().split())


def txt(v):
    return "" if v is None else str(v).strip()


def iso(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    raise ValueError("data nao reconhecida / unparsable date: %r" % (v,))


def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    if not os.path.exists(xlsx):
        sys.exit("Planilha nao encontrada / spreadsheet not found:\n  %s" % xlsx)

    print("Lendo / Reading: %s" % xlsx)
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit("Aba '%s' nao existe. Abas: %s" % (SHEET, wb.sheetnames))
    ws = wb[SHEET]
    rows = ws.iter_rows(values_only=True)

    header = next(rows)
    want = {norm(k): v for k, v in COLS.items()}
    idx = {}
    for i, h in enumerate(header):
        key = want.get(norm(h))
        if key and key not in idx:
            idx[key] = i

    missing = sorted(set(COLS.values()) - set(idx))
    if missing:
        sys.exit(
            "Colunas ausentes na planilha / missing columns:\n  %s\n"
            "Cabecalhos encontrados / headers found:\n  %s"
            % (", ".join(missing), [txt(h) for h in header if h])
        )

    def cell(row, key):
        i = idx[key]
        return row[i] if i < len(row) else None

    entries, seen_ids, dupes, no_id = [], set(), [], []
    for row in rows:
        if not any(v not in (None, "") for v in row):
            continue
        num = cell(row, "num")
        if num in (None, ""):
            continue
        eid = txt(cell(row, "id"))
        name_pt = txt(cell(row, "namePt"))
        if not eid:
            no_id.append(name_pt)
            continue
        if eid in seen_ids:
            dupes.append((eid, name_pt))
        seen_ids.add(eid)

        flags = [
            k[5:] for k in COLS.values()
            if k.startswith("flag_") and txt(cell(row, k))
        ]
        debut = {k: iso(cell(row, "d_" + k)) for k in DATE_KEYS}

        entries.append({
            "id": eid,
            "num": int(num),
            "namePt": name_pt,
            "nameEn": txt(cell(row, "nameEn")) or name_pt,
            "speciesPt": txt(cell(row, "speciesPt")),
            "speciesEn": txt(cell(row, "speciesEn")),
            "region": txt(cell(row, "region")) or "Indefinida",
            "regFormPt": txt(cell(row, "regFormPt")),
            "regFormEn": txt(cell(row, "regFormEn")),
            "altFormPt": txt(cell(row, "altFormPt")),
            "altFormEn": txt(cell(row, "altFormEn")),
            "costumePt": txt(cell(row, "costumePt")),
            "costumeEn": txt(cell(row, "costumeEn")),
            "gender": txt(cell(row, "gender")),
            "flags": flags,
            # Legado: linha "canonica" da planilha. Nao usar em contagens.
            # Legacy canonical-row marker. Never drive counts with this.
            "dexCanon": bool(txt(cell(row, "dexCanon"))),
            "debut": debut,
        })

    entries.sort(key=lambda e: (e["num"], e["id"]))

    # Ordem canonica das regioes; o que nao estiver na lista vai para o fim.
    REGION_ORDER = ["Kanto", "Johto", "Hoenn", "Sinnoh", "Unova", "Kalos",
                    "Alola", "Galar", "Hisui", "Paldea"]
    found = {e["region"] for e in entries}
    regions = [r for r in REGION_ORDER if r in found]
    regions += sorted(r for r in found if r not in REGION_ORDER)

    out = {
        "generated": datetime.now().replace(microsecond=0).isoformat(),
        "source": os.path.basename(xlsx),
        "spriteBase": "https://raw.githubusercontent.com/Gabrielense/"
                      "pogorewind/main/sprites/",
        "markKeys": MARK_KEYS,
        "dateKeys": DATE_KEYS,
        "regions": regions,
        "stats": {
            "entries": len(entries),
            "dexNumbers": len({e["num"] for e in entries}),
            "released": sum(1 for e in entries if e["debut"]["base"]),
        },
        # avisos mostrados na tela "Meus dados" do site
        "warnings": [],
        "entries": entries,
    }

    # ------------------------------------------------------ sanity checks
    problems = []
    if len(entries) < 1500:
        problems.append("poucas entradas (%d) / too few entries" % len(entries))
    if dupes:
        problems.append(
            "IDs duplicados / duplicate IDs (%d): %s"
            % (len(dupes), ", ".join("%s=%s" % d for d in dupes[:8]))
        )
    if no_id:
        problems.append(
            "linhas sem ID / rows without ID (%d): %s"
            % (len(no_id), ", ".join(no_id[:8]))
        )
    bad_dates = [
        e["id"] for e in entries
        if e["debut"]["shiny"] and not e["debut"]["base"]
    ]
    if bad_dates:
        problems.append(
            "estreia brilhante sem estreia base (%d): %s"
            % (len(bad_dates), ", ".join(bad_dates[:8]))
        )

    # O ID comeca com o numero da dex. Se o Numero da linha nao bate, quase
    # sempre e autofill do Excel arrastando a coluna Numero. Ja aconteceu com
    # 8 linhas de Pikachu (26,27,...,32 em vez de 25).
    mismatched = []
    for e in entries:
        head = e["id"].split("-")[0].split("+")[0].rstrip("MG")
        if head.isdigit() and int(head) != e["num"]:
            mismatched.append((e["namePt"], e["num"], int(head), e["id"]))
    if mismatched:
        problems.append(
            "Numero nao bate com o ID (%d) - provavel autofill do Excel:\n      %s"
            % (len(mismatched),
               "\n      ".join("%-34s Numero=%-4s deveria ser %-4s (ID %s)"
                               % (n, a, b, i) for n, a, b, i in mismatched[:12]))
        )
        out["warnings"].append({
            "what": "%d linhas com o Número diferente do ID" % len(mismatched),
            "why": "Provável autofill do Excel arrastando a coluna Número. "
                   "Corrija para o número do ID: "
                   + "; ".join("%s -> %d" % (n, b) for n, a, b, i in mismatched[:10]),
        })

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as fh:
        json.dump(out, fh, ensure_ascii=False, separators=(",", ":"))

    size_kb = os.path.getsize(OUT_PATH) / 1024
    print("OK -> %s (%.0f KB)" % (OUT_PATH, size_kb))
    print("   entradas/entries: %d | dex: %d | lancadas/released: %d"
          % (out["stats"]["entries"], out["stats"]["dexNumbers"],
             out["stats"]["released"]))
    if problems:
        print("\nAVISOS / WARNINGS:")
        for p in problems:
            print("  ! " + p)
        # Duplicatas e linhas sem ID quebram o casamento na importacao.
        if dupes or no_id:
            sys.exit(1)


if __name__ == "__main__":
    main()
