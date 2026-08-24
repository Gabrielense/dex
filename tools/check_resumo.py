#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Confere a agregacao do site contra a aba Resumo da planilha.
Reconciles the site's aggregation against the spreadsheet's Resumo sheet.

Dois modos / two modes:
  legacy : filtra Dex="Sim" (uma linha canonica por numero) -> DEVE bater 100%
           com o Resumo, provando que a leitura dos dados esta correta.
  anyEntry : agrupa por Numero e aceita qualquer entrada (regra correta do jogo)
           -> pode divergir do Resumo; as diferencas sao o ganho da correcao.

Uso / Usage:
    python tools/check_resumo.py [caminho/para/PokeAgenda.xlsx]
"""

import json
import os
import sys
from datetime import date, datetime

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl nao encontrado. Rode: pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DEFAULT_XLSX = os.path.abspath(os.path.join(ROOT, "..", "PokéAgenda 2026.xlsx"))
SHEET_MAIN, SHEET_RESUMO = "PokéAgenda", "Resumo"
RESUMO_REGION_ROWS = range(2, 13)   # Kanto..Indefinida
RESUMO_REGION_COL = 13              # coluna M

# A aba Resumo guarda os valores que o Excel calculou da ultima vez que o
# arquivo foi aberto - ou seja, com o TODAY() daquele dia. Se hoje for depois
# disso e algo estreou no meio, a comparacao acusa diferenca que nao existe.
# Por isso o verificador tenta hoje e, se nao bater, volta alguns dias ate
# achar a data em que a planilha foi recalculada.
TODAY = date.today()
LOOKBACK_DAYS = 10

sys.path.insert(0, HERE)
from export_skeleton import COLS, DATE_KEYS, norm, txt, iso  # noqa: E402


# --------------------------------------------------------------- load rows
def load_rows(xlsx):
    """Le a aba 1 inteira: esqueleto + marcas pessoais (so para conferir)."""
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    ws = wb[SHEET_MAIN]
    it = ws.iter_rows(values_only=True)
    header = next(it)
    want = {norm(k): v for k, v in COLS.items()}
    idx = {}
    for i, h in enumerate(header):
        k = want.get(norm(h))
        if k and k not in idx:
            idx[k] = i

    rows = []
    for row in it:
        num = row[idx["num"]] if idx["num"] < len(row) else None
        if num in (None, ""):
            continue

        def g(key):
            i = idx[key]
            return row[i] if i < len(row) else None

        rows.append({
            "id": txt(g("id")),
            "num": int(num),
            "namePt": txt(g("namePt")),
            "speciesPt": txt(g("speciesPt")),
            "region": txt(g("region")) or "Indefinida",
            "regFormPt": txt(g("regFormPt")),
            "altFormPt": txt(g("altFormPt")),
            "costumePt": txt(g("costumePt")),
            "dexCanon": bool(txt(g("dexCanon"))),
            "flags": {
                k[5:] for k in COLS.values()
                if k.startswith("flag_") and txt(g(k))
            },
            "debut": {k: iso(g("d_" + k)) for k in DATE_KEYS},
            "marks": {
                m: bool(txt(g(m))) for m in
                ["caught", "shiny", "shadow", "purified", "shadowShiny",
                 "dmax", "dmaxShiny", "shinyDex", "xxs", "xxl", "lucky",
                 "perfect"]
            },
        })
    return rows


AS_OF = TODAY   # data usada nas contas; ajustada em main()


def released(entry, gate):
    d = entry["debut"].get(gate)
    return bool(d) and date.fromisoformat(d) <= AS_OF


def match_cond(entry, cond):
    if "flag" in cond:
        return cond["flag"] in entry["flags"]
    if "nonempty" in cond:
        return bool(entry[cond["nonempty"]])
    if "equals" in cond:
        eq = cond["equals"]
        return entry.get(eq["field"], "") == eq["value"]
    return True


def in_subset(entry, cat):
    sub = cat.get("subset")
    if not sub:
        return True
    if "any" in sub:
        return any(match_cond(entry, c) for c in sub["any"])
    return match_cond(entry, sub)


def dex_region_map(rows):
    """Regiao de cada numero da dex = a da entrada base (id de 4 digitos)."""
    out = {}
    for e in rows:
        if e["id"] == "%04d" % e["num"]:
            out[e["num"]] = e["region"]
    for e in rows:
        if e["num"] not in out and e["dexCanon"]:
            out[e["num"]] = e["region"]
    for e in rows:
        out.setdefault(e["num"], e["region"])
    return out


# ------------------------------------------------------------- aggregation
def aggregate(rows, cat, mode):
    """-> {regiao: {'caught':n,'released':n,'total':n}}"""
    acc = {}

    def bucket(region):
        return acc.setdefault(
            region, {"caught": 0, "released": 0, "total": 0})

    gate, mark = cat["gate"], cat["mark"]

    if cat["scope"] == "entry":
        for e in rows:
            if not in_subset(e, cat):
                continue
            b = bucket(e["region"])
            b["total"] += 1
            if released(e, gate):
                b["released"] += 1
                if e["marks"][mark]:
                    b["caught"] += 1
        return acc

    # scope == "dex"
    if mode == "legacy":
        for e in rows:
            if not e["dexCanon"] or not in_subset(e, cat):
                continue
            b = bucket(e["region"])
            b["total"] += 1
            if released(e, gate):
                if cat.get("denomExcludeFlag") in e["flags"]:
                    pass  # fora do denominador (ex.: nao trocavel)
                else:
                    b["released"] += 1
                if e["marks"][mark]:
                    b["caught"] += 1
        return acc

    # mode == "anyEntry": agrupa por numero, qualquer entrada conta
    regions = dex_region_map(rows)
    groups = {}
    for e in rows:
        if in_subset(e, cat):
            groups.setdefault(e["num"], []).append(e)
    for num, group in groups.items():
        b = bucket(regions[num])
        b["total"] += 1
        rel = [e for e in group if released(e, gate)]
        in_denom = bool(rel)
        excl = cat.get("denomExcludeFlag")
        if in_denom and excl:
            in_denom = any(excl not in e["flags"] for e in rel)
        if in_denom:
            b["released"] += 1
            # "pego" so conta dentro do que ja foi lancado, senao pegos > lancados
            if any(e["marks"][mark] for e in group):
                b["caught"] += 1
    return acc


# ------------------------------------------------------------------ report
def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    if not os.path.exists(xlsx):
        sys.exit("Planilha nao encontrada:\n  %s" % xlsx)

    with open(os.path.join(ROOT, "data", "categories.json"), encoding="utf-8") as fh:
        catdoc = json.load(fh)
    cats = catdoc["categories"]
    gaps = catdoc.get("_knownSheetGaps", {}).get("gaps", [])
    stale = catdoc.get("_staleUntilRecalc", {}).get("gaps", [])

    def _match(rules, key, region, field):
        return any(
            g["category"] == key and region in g["regions"]
            and field in g["fields"] for g in rules
        )

    def is_known_gap(key, region, field):
        return _match(gaps, key, region, field)

    def is_stale(key, region, field):
        return _match(stale, key, region, field)

    rows = load_rows(xlsx)
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[SHEET_RESUMO]
    sheet_regions = {}
    for r in RESUMO_REGION_ROWS:
        name = txt(ws.cell(r, RESUMO_REGION_COL).value)
        if name:
            sheet_regions[name] = r

    # Descobre com que data a aba Resumo foi calculada: tenta hoje e vai
    # voltando ate os numeros baterem. Sem isso, virar a meia-noite ja faz o
    # verificador acusar erro que nao existe.
    global AS_OF
    from datetime import timedelta

    def count_fails(when):
        global AS_OF
        AS_OF = when
        n = 0
        for cat in cats:
            legacy = aggregate(rows, cat, "legacy")
            for region, srow in sheet_regions.items():
                for field, col in cat.get("resumo", {}).items():
                    want = ws.cell(srow, col).value
                    want = 0 if want in (None, "") else int(want)
                    if legacy.get(region, {}).get(field, 0) != want \
                            and not is_known_gap(cat["key"], region, field) \
                            and not is_stale(cat["key"], region, field):
                        n += 1
        return n

    best, best_fails = TODAY, count_fails(TODAY)
    if best_fails:
        for back in range(1, LOOKBACK_DAYS + 1):
            cand = TODAY - timedelta(days=back)
            n = count_fails(cand)
            if n < best_fails:
                best, best_fails = cand, n
            if n == 0:
                break
    AS_OF = best

    print("Planilha / spreadsheet : %s" % os.path.basename(xlsx))
    print("Entradas / entries     : %d" % len(rows))
    print("Data de hoje / today   : %s" % TODAY.isoformat())
    if best != TODAY:
        print("Resumo calculado em    : %s  <- a planilha nao e reaberta no Excel"
              % best.isoformat())
        print("                         desde entao; comparando com essa data.")
    print("Regioes no Resumo      : %s\n" % ", ".join(sheet_regions))

    fails, diffs, fixed, pending = [], [], [], []
    for cat in cats:
        legacy = aggregate(rows, cat, "legacy")
        anyent = aggregate(rows, cat, "anyEntry")
        rmap = cat.get("resumo", {})

        for region, srow in sheet_regions.items():
            for field in ("caught", "released", "total"):
                col = rmap.get(field)
                if not col:
                    continue
                want = ws.cell(srow, col).value
                want = 0 if want in (None, "") else int(want)
                got = legacy.get(region, {}).get(field, 0)
                if got == want:
                    continue
                if is_known_gap(cat["key"], region, field):
                    fixed.append((cat["key"], region, field, want, got))
                elif is_stale(cat["key"], region, field):
                    pending.append((cat["key"], region, field, want, got))
                else:
                    fails.append((cat["key"], region, field, want, got))

        lc = sum(v["caught"] for v in legacy.values())
        ac = sum(v["caught"] for v in anyent.values())
        lr = sum(v["released"] for v in legacy.values())
        ar = sum(v["released"] for v in anyent.values())
        if cat["scope"] == "dex" and (lc, lr) != (ac, ar):
            diffs.append((cat["key"], lc, ac, lr, ar))

    print("=" * 68)
    print("1) MODO LEGADO vs RESUMO  (tem que bater 100%)")
    print("   LEGACY MODE vs RESUMO  (must match 100%)")
    print("=" * 68)
    if fails:
        print("  %d divergencia(s) / mismatch(es):" % len(fails))
        for key, region, field, want, got in fails[:40]:
            print("   x %-20s %-12s %-9s planilha=%-5s calculado=%s"
                  % (key, region, field, want, got))
        if len(fails) > 40:
            print("   ... e mais %d" % (len(fails) - 40))
    else:
        print("  OK - todas as celulas conferem / every cell matches.")

    if fixed:
        print("\n  Falhas conhecidas da planilha, ja corrigidas pelo site:")
        print("  Known spreadsheet gaps, already fixed by the site:")
        for key, region, field, want, got in fixed:
            print("   + %-20s %-12s %-9s planilha=%-5s site=%s"
                  % (key, region, field, want, got))

    if pending:
        print("\n  Aguardando recalculo: o Resumo em cache e de antes dos")
        print("  consertos. Abra a planilha no Excel e salve; estas somem.")
        for key, region, field, want, got in pending:
            print("   ~ %-20s %-12s %-9s cache=%-5s correto=%s"
                  % (key, region, field, want, got))

    print("\n" + "=" * 68)
    print("2) CORRECAO 'qualquer entrada conta' (regra real do jogo)")
    print("   'ANY ENTRY COUNTS' FIX (the game's real rule)")
    print("=" * 68)
    if diffs:
        print("  %-22s %-17s %s" % ("categoria", "pegos leg->novo", "lancados leg->novo"))
        for key, lc, ac, lr, ar in diffs:
            print("  %-22s %5d -> %-9d %5d -> %d"
                  % (key, lc, ac, lr, ar))
        print("\n  Diferenca = numeros da dex que a planilha nao contava porque")
        print("  a linha canonica nao representava o que voce tem.")
    else:
        print("  Nenhuma diferenca / no differences.")

    print("")
    if fails:
        sys.exit(1)


if __name__ == "__main__":
    main()
